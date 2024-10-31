import openpyxl
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template
from flask_mail import Mail, Message
import uuid
import os

from pathlib import Path
# from dotenv import load_dotenv

from get_data import get_data


# current_dir = Path(__file__).resolve().parent if "__file__" in locals() else Path.cwd()
# envars = current_dir / ".env"
# load_dotenv(envars)

app = Flask(__name__)
#CORS(app)
app.secret_key = str(uuid.uuid4())[:50]
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = ''# Your Email
app.config['MAIL_PASSWORD'] = ''# Your app Password

mail = Mail(app)

# def get_intern_info():
#     """
#     get_intern_info: cleaning up and fetching intern's datas from the spreadsheet
#     Returns:
#         A list for dictionary with key values of the inters data
#     """
#     wb = load_workbook(filename='AI Talent.xlsx', data_only=True)
#     ws = wb.active

#     names = [cell.value for row in ws.iter_rows(min_row=2, max_col= 1, max_row=len(tuple(ws.rows))) for cell in row]
#     emails = [cell.value for cell in ws['C'] if cell.value != 'email']
#     total_score = [cell.value for cell in ws['J'] if cell.value != 'Total']
#     interns = []

#     for i in range(len(tuple(ws.rows)) - 1):
#         intern = {"name": names[i], "email": emails[i], "total_score": round(total_score[i],2)}
#         interns.append(intern)

#     # print(interns)
#     return interns

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/send_email')
def SendEmail():
    """
    SendEmail route is tasked with sending email vial flasks mail library
    when the /send_email route is triggerd
    Returns:
        An alert message showing the success
    """
    # interns = get_intern_info()
    interns = get_data()
    for i in interns:
         msg = Message('Nova ',
                       sender="Noreply@gmail.com",
                       recipients=[i["email"]])
         #msg.body = f"Dear {i['name']} This email was sent from Ic_labs."
         msg.html = f"""
         <h1>Dear {i['name']} This email was sent from Icog Labs</h1>
         <h3>Your Score for the week is</h3><p>{i['results']}.</p>"""
         mail.send(msg)
         print(msg)


    #test the app by using some valid email
    """
    msg = Message('Nova ',
                       sender='NoReply@gmail.com',
                       recipients=[''])
    msg.body = f"Dear someone This email was sent from Ic_labs."
    msg.html = f"<h1>Your Score for the week is</h1><p>.</p>"
    mail.send(msg)
    print(msg)
    """
    return render_template('alert.html')
 

if __name__ == "__main__":
    app.run(debug=True)
